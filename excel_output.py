import openpyxl

import logger


class ExcelOutput:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def find_day_column(self, day_num):
        for col in self.sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                logger.LOGGER().debug(f"CELL ADDRESS: {cell.column_letter+"1"}, VALUE: {cell.value} - Looking for: {day_num}")
                if cell.value == day_num:
                    return cell.column
        return None

    def write_blob_counts(self, day_num, sample_number, num_keypoints):
        col = self.find_day_column(day_num)
        if col is None:
            raise ValueError(f"Day {day_num} not found in the first row of the Excel sheet.")

        # Find the "colonies" cell under the day_num column
        for row in self.sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if str(cell.value).lower() == "colonies":
                    colonies_row = cell.row
                    break
            else:
                continue
            break
        else:
            raise ValueError(f"'colonies' cell not found under day {day_num} column.")

        # Write the number of keypoints under the "colonies" cell
        target_cell = self.sheet.cell(row=colonies_row + sample_number, column=col)
        target_cell.value = num_keypoints

    def save(self):
        self.workbook.save(self.file_path)